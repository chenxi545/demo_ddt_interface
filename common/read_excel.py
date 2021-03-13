import openpyxl

class Case:
    def __init__(self, attrs):
        # 这个类用来存储用例
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """读取数据"""

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name: 文件名字 --> str
        :param sheet_name: 表单名字 --> str
        """
        self.file_name = file_name
        # 打开工作薄
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[sheet_name]

    def read_data_line(self):
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for case in rows_data[1:]:
            # 获取一条测试用例
            data = []
            for cell in case:
                if isinstance(cell.value, str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            # 将该条数据放入cases中
            case_data = dict(list(zip(titles, data)))
            cases.append(case_data)
        return cases

    def read_data_obj(self):
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for case in rows_data[1:]:
            # 获取一条测试用例
            case_obj = case()
            data = []
            for cell in case:
                if isinstance(cell.value, str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))

            for i in case_obj:
                setattr(case_obj, i[0], i[1])

            # print(case_obj.id, case_obj.data)

        return cases

    def r_data(self, list1):
        # 获取最大行数
        max_r = self.sheet.max_row
        # 空列表存放所有用例
        cases = []
        # 空列表存放表头
        titles = []
        # 遍历所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                    case = dict(list(zip(titles, case_data)))
                    cases.append(case)
            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases

    def r_data_obj(self, list1):
        # 获取最大行数
        max_r = self.sheet.max_row
        # 空列表存放所有用例
        cases = []
        # 空列表存放表头
        titles = []
        # 遍历所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row != 1:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                    case = dict(list(zip(titles, case_data)))
                    cases.append(case)
            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases

    def write_data(self, row, column, msg):
        self.sheet.cell(row=row, column=column, msg=msg)
        self.wb.save(self.file_name)
